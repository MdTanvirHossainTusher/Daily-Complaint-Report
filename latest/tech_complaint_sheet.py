import openpyxl
from datetime import datetime, timedelta

daily_technology_file_name = r'Daily Technology Complaint Report2.xlsx'
raw_dump = r'Raw Dump.xlsx'
tech_complaint_sheet_name = 'Tech complaint'
raw_pivot_sheet_name = 'RAW Pivot'
assign_pivot_sheet_name = 'Assigned Pivot'


def load_workbook(daily_technology_file_name, raw_dump):
    technology_wb = openpyxl.load_workbook(daily_technology_file_name)
    raw_dump_wb = openpyxl.load_workbook(raw_dump)
    return technology_wb, raw_dump_wb


def load_worksheet(technology_wb, raw_dump_wb):
    technology_sheet = technology_wb[tech_complaint_sheet_name]
    raw_pivot_sheet = raw_dump_wb[raw_pivot_sheet_name]
    assign_pivot_sheet = raw_dump_wb[assign_pivot_sheet_name]
    return technology_sheet, raw_pivot_sheet, assign_pivot_sheet


def insert_dates_to_list(assign_from_date, assign_to_date):
    # Convert strings to datetime objects
    from_date_obj = datetime.strptime(assign_from_date, "%d-%b-%y")
    to_date_obj = datetime.strptime(assign_to_date, "%d-%b-%y")
    current_year = from_date_obj.year

    # Generate a list of date strings
    date_list = [from_date_obj + timedelta(days=i) for i in range((to_date_obj - from_date_obj).days + 1)]
    formatted_date_list = [date.strftime("%d-%b") for date in date_list]
    # print(current_year, formatted_date_list, end=" ----------------------\n ")
    return current_year, formatted_date_list


# Convert target dates to 'YYYY-MM-DD' format
def format_inserted_dates(target_year, target_dates):
    formatted_dates = []
    for target_date in target_dates:
        try:
            date_obj = datetime.strptime(f"{target_date}-{target_year}", '%d-%b-%Y')
            formatted_date = date_obj.strftime('%Y-%m-%d')
            formatted_dates.append(formatted_date)
        except ValueError:
            print(f"Invalid date format: {target_date}. Please use the '2-Dec' format.")
    return formatted_dates


def get_column_indices(technology_sheet, formatted_dates):
    column_indices = []
    # Iterate through target dates and find the column indices
    for target_date in formatted_dates:
        column_index = None
        for col_num in range(1, technology_sheet.max_column + 1):
            if str(technology_sheet.cell(row=1, column=col_num).value).split(' ')[0] == target_date:
                column_index = col_num
                column_indices.append(column_index)
                break

        # Check if the column was found
        if column_index is not None:
            print(f"The column index for '{target_date}' is: {column_index}")
        else:
            print(f"Column '{target_date}' not found in the sheet.")
    return column_indices


def fill_tech_complaint_columns(technology_sheet, assign_pivot_sheet, column_indices):
    # assign pivot theke value ene bosaitesi
    for i in technology_sheet.iter_rows():  # iter_rows() --> 0 based index

        complain_name = i[0].value  # 7 = Team, 8 = sub category, heading row = 1
        row_number = i[0].row

        for j in assign_pivot_sheet.iter_rows():
            if complain_name is not None and j[0].value is not None and j[0].value.lower() == complain_name.lower():
                for k in range(len(column_indices)):
                    technology_sheet.cell(row=row_number, column=column_indices[k]).value = j[k+1].value


def fill_blank_with_na(technology_sheet, column_indices):
    # sei ghor gula #N/A or None segulare 0 diye fillup kortesi
    for row in technology_sheet.iter_rows():
        complain_name = row[0].value
        row_number = row[0].row

        if complain_name is None or complain_name == ' ' or complain_name == '':
            continue

        for k in range(len(column_indices)):  # 5, 6
            if technology_sheet.cell(row=row_number, column=column_indices[k]).value is None:  # 1 based index
                technology_sheet.cell(row=row_number, column=column_indices[k]).value = 0


def fill_ccd_technology_fields(technology_sheet, raw_pivot_sheet, assign_pivot_sheet, column_indices):
    # Open at CCD, Assign to Technology ghor fillup kortesi
    raw_assign_pivot_rows = [5, 12, 19]
    p = 0
    q = 0
    for row in technology_sheet.iter_rows():
        complain_name = row[0].value
        row_number = row[0].row

        if complain_name is not None and complain_name.strip().lower() == 'Open at CCD'.lower():
            starting_col = 'B'
            for k in range(len(column_indices)):  # column_indices = [14, 15]
                raw_pivot_cell_value = f"{starting_col}{raw_assign_pivot_rows[p]}"  # B5, C5
                starting_col = chr(ord(starting_col) + 1)  # C
                technology_sheet.cell(row=row_number, column=column_indices[k]).value = raw_pivot_sheet[raw_pivot_cell_value].value
            p += 1

        elif complain_name is not None and complain_name.strip().lower() == 'Assign to Technology'.lower():
            starting_col = 'B'
            for r in range(len(column_indices)):
                assign_pivot_cell_value = f"{starting_col}{raw_assign_pivot_rows[q]}"  # B5, C5
                starting_col = chr(ord(starting_col) + 1)  # C
                technology_sheet.cell(row=row_number, column=column_indices[r]).value = assign_pivot_sheet[assign_pivot_cell_value].value
            q += 1
    column_indices.clear()


def save_file(technology_wb):
    technology_wb.save(daily_technology_file_name)
    # technology_wb.save(r'Daily Technology Complaint Report3.xlsx')

# Close the workbook
# technology_wb.close()


def daily_technology_tech_complaint_sheet_processing(assign_from_date, assign_to_date):
    technology_wb, raw_dump_wb = load_workbook(daily_technology_file_name, raw_dump)
    technology_sheet, raw_pivot_sheet, assign_pivot_sheet = load_worksheet(technology_wb, raw_dump_wb)
    current_year, formatted_date_list = insert_dates_to_list(assign_from_date, assign_to_date)
    formatted_dates = format_inserted_dates(current_year, formatted_date_list)
    column_indices = get_column_indices(technology_sheet, formatted_dates)
    fill_tech_complaint_columns(technology_sheet, assign_pivot_sheet, column_indices)
    fill_blank_with_na(technology_sheet, column_indices)
    fill_ccd_technology_fields(technology_sheet, raw_pivot_sheet, assign_pivot_sheet, column_indices)
    save_file(technology_wb)
