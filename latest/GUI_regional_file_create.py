import openpyxl
import tkinter as tk
from tkinter import filedialog
from tkcalendar import DateEntry

daily_dump_file_name_raw_dump = r'Daily_Dump(Updated).xlsx'
team_col_index = 8
hash_na = '#N/A'
output_file = r"regional.xlsx"
# assign_from_date = ""
# assign_to_date = ""
open_date_col_index = 20
assign_date_col_index = 21
assign_date_filtered_rows = []


def load_workbook(file_path):
    return openpyxl.load_workbook(file_path)


def load_worksheet(daily_dump):
    return daily_dump.active


def create_regional_file(output_path):
    regional_workbook = openpyxl.Workbook()
    default_sheet = regional_workbook['Sheet']
    regional_workbook.remove(default_sheet)
    regional_sheet = regional_workbook.create_sheet("Sheet1")
    regional_workbook.save(output_path)
    return regional_workbook, regional_sheet


def select_prv_assigned_date(daily_dump_sheet, assign_from_date, assign_to_date):
    heading_row = next(
        daily_dump_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assign_date_filtered_rows = [heading_row]

    for row in daily_dump_sheet.iter_rows(min_row=1, values_only=True):
        team_value = row[team_col_index - 1]
        assign_date_value = row[assign_date_col_index]

        if team_value != hash_na and assign_date_value is not None and str(assign_from_date).lower() <= str(
                assign_date_value).lower() <= str(assign_to_date).lower():
            assign_date_filtered_rows.append(row)

    return assign_date_filtered_rows


def paste_assigned_date_to_assigned_sheet(assigned_sheet_of_raw_dump, assign_date_filtered_rows):
    for row_data in assign_date_filtered_rows:
        assigned_sheet_of_raw_dump.append(row_data)


def save_regional_file(regional, output_path):
    regional.save(output_path)


# def browse_file():
#     global daily_dump_file_name_raw_dump
#     daily_dump_file_name_raw_dump = filedialog.askopenfilename(
#         title="Select Daily Dump File",
#         filetypes=[("Excel Files", "*.xlsx"), ("CSV files", "*.csv")]
#     )
#     print(f"Selected File: {daily_dump_file_name_raw_dump}")


# def browse_output_directory():
#     global output_file
#     output_file = filedialog.asksaveasfilename(
#         title="Save Regional File As",
#         filetypes=[("Excel Files", "*.xlsx")]
#     )
#     print(f"Output File: {output_file}")


# def browse_dates():
    # global assign_from_date, assign_to_date
    # assign_from_date = entry_from_date.get()
    # assign_to_date = entry_to_date.get()
    # print(f"Assign From Date: {assign_from_date}")
    # print(f"Assign To Date: {assign_to_date}")

    # global assign_from_date, assign_to_date
    # assign_from_date = entry_from_date.get_date().strftime("%d-%b-%y")
    # assign_to_date = entry_to_date.get_date().strftime("%d-%b-%y")
    # print(f"Assign From Date: {assign_from_date}")
    # print(f"Assign To Date: {assign_to_date}")


def regional_file_creation_process(assign_from_date, assign_to_date):
    daily_dump = load_workbook(daily_dump_file_name_raw_dump)
    daily_dump_sheet = load_worksheet(daily_dump)
    regional_workbook, regional_sheet = create_regional_file(output_file)

    assign_date_filtered_rows = select_prv_assigned_date(daily_dump_sheet, assign_from_date, assign_to_date)
    paste_assigned_date_to_assigned_sheet(regional_sheet, assign_date_filtered_rows)

    save_regional_file(regional_workbook, output_file)
    assign_date_filtered_rows.clear()

    # root.destroy()  # Close the GUI window

'''
root = tk.Tk()
root.title("Regional File Generator")

root.geometry("400x300")  # width and height

# File Selection
btn_browse_file = tk.Button(root, text="Browse Daily Dump File", command=browse_file)
btn_browse_file.pack(pady=10)

# # Output Directory Selection
# btn_browse_output = tk.Button(root, text="Select Output Directory", command=browse_output_directory)
# btn_browse_output.pack(pady=10)


# Date Selection using tkcalendar
entry_from_date = DateEntry(root, width=15, justify="center")
entry_to_date = DateEntry(root, width=15, justify="center")

# Set the date pattern directly on the DateEntry widget
entry_from_date.date_pattern = "%d-%b-%y"
entry_to_date.date_pattern = "%d-%b-%y"

entry_from_date.pack(pady=5)
entry_to_date.pack(pady=5)

btn_browse_dates = tk.Button(root, text="Apply Dates", command=browse_dates)
btn_browse_dates.pack(pady=10)

# Run Process Button
btn_run_process = tk.Button(root, text="Run Process", command=run_process)
btn_run_process.pack(pady=20)

root.mainloop()
'''