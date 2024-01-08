import openpyxl
from datetime import datetime

daily_dump_file_name_raw_dump = r'Daily_Dump(Updated).xlsx'
team_col_index = 8
hash_na = '#N/A'
output_file = r"regional.xlsx"
open_date_col_index = 20
assign_date_col_index = 21
assign_date_filtered_rows = []


def load_workbook(file_path):
    return openpyxl.load_workbook(file_path)


def load_worksheet(daily_dump):
    return daily_dump.active


def create_regional_file(output_path):
    regional_workbook = openpyxl.Workbook()
    default_sheet = regional_workbook['Sheet']
    regional_workbook.remove(default_sheet)
    regional_sheet = regional_workbook.create_sheet("Sheet1")
    regional_workbook.save(output_path)
    return regional_workbook, regional_sheet


def select_prv_assigned_date(daily_dump_sheet, assign_from_date, assign_to_date):
    heading_row = next(
        daily_dump_sheet.iter_rows(min_row=1, max_row=1, values_only=True))

    assign_date_filtered_rows = [heading_row]

    # Convert string dates to datetime objects
    assign_from_date = datetime.strptime(assign_from_date, "%d-%b-%y")
    assign_to_date = datetime.strptime(assign_to_date, "%d-%b-%y")

    for row in daily_dump_sheet.iter_rows(min_row=1, values_only=True):  # tuple = iter_rows() --> 0 based index

        team_value = row[team_col_index - 1]
        assign_date_string = row[assign_date_col_index]

        if assign_date_string == 'ASSIGNED_DATE':
            continue

        assign_date_value = datetime.strptime(assign_date_string, "%d-%b-%y") if assign_date_string else None

        if team_value != hash_na and assign_date_value is not None and assign_from_date <= assign_date_value <= assign_to_date:
            assign_date_filtered_rows.append(row)

    print(len(assign_date_filtered_rows), end=" Assigned Date rows\n")
    return assign_date_filtered_rows


def paste_assigned_date_to_assigned_sheet(assigned_sheet_of_raw_dump, assign_date_filtered_rows):
    for row_data in assign_date_filtered_rows:
        assigned_sheet_of_raw_dump.append(row_data)


def save_regional_file(regional, output_path):
    regional.save(output_path)


def regional_file_creation_process(assign_from_date, assign_to_date):
    daily_dump = load_workbook(daily_dump_file_name_raw_dump)
    daily_dump_sheet = load_worksheet(daily_dump)
    regional_workbook, regional_sheet = create_regional_file(output_file)

    assign_date_filtered_rows = select_prv_assigned_date(daily_dump_sheet, assign_from_date, assign_to_date)
    paste_assigned_date_to_assigned_sheet(regional_sheet, assign_date_filtered_rows)

    save_regional_file(regional_workbook, output_file)
    assign_date_filtered_rows.clear()
