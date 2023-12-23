import openpyxl
from datetime import datetime, timedelta

daily_technology_file_name = r'Daily Technology Complaint Report.xlsx'
regional_file_name = r'regional.xlsx'
radio_voice_sheet_name = 'Region wise Radio-Voice'
radio_data_sheet_name = 'Region wise Radio-Data'
voice_sheet_name = "voice_pivot"
data_sheet_name = "data_pivot"


def load_workbook(daily_technology_file_name, regional_file_name):
    technology_region_wb = openpyxl.load_workbook(daily_technology_file_name)
    regional_wb = openpyxl.load_workbook(regional_file_name)
    return technology_region_wb, regional_wb


def load_worksheet(technology_region_wb, regional_wb):
    region_wise_radio_voice_sheet = technology_region_wb[radio_voice_sheet_name]
    region_wise_radio_data_sheet = technology_region_wb[radio_data_sheet_name]
    voice_regional_pivot_sheet = regional_wb[voice_sheet_name]
    data_regional_pivot_sheet = regional_wb[data_sheet_name]
    return region_wise_radio_voice_sheet, region_wise_radio_data_sheet, voice_regional_pivot_sheet, data_regional_pivot_sheet


def insert_dates_to_list(assign_from_date, assign_to_date):

    # Convert strings to datetime objects
    from_date_obj = datetime.strptime(assign_from_date, "%d-%b-%y")
    to_date_obj = datetime.strptime(assign_to_date, "%d-%b-%y")
    current_year = from_date_obj.year

    # Generate a list of date strings
    date_list = [from_date_obj + timedelta(days=i) for i in range((to_date_obj - from_date_obj).days + 1)]
    formatted_date_list = [date.strftime("%d-%b") for date in date_list]
    return current_year, formatted_date_list


def format_inserted_dates(target_year, target_dates):
    # Convert target dates to 'YYYY-MM-DD' format
    formatted_dates = []
    for target_date in target_dates:
        try:
            date_obj = datetime.strptime(f"{target_date}-{target_year}", '%d-%b-%Y')
            formatted_date = date_obj.strftime('%Y-%m-%d')
            formatted_dates.append(formatted_date)
        except ValueError:
            print(f"Invalid date format: {target_date}. Please use the '2-Dec' format.")
    return formatted_dates


def get_column_indices(region_wise_radio_voice_sheet, formatted_dates):
    column_indices = []
    # Iterate through target dates and find the column indices
    for target_date in formatted_dates:
        column_index = None
        for col_num in range(1, region_wise_radio_voice_sheet.max_column + 1):
            # date time --> just take the date
            if str(region_wise_radio_voice_sheet.cell(row=2, column=col_num).value).split(' ')[0] == target_date:
                column_index = col_num
                column_indices.append(column_index)
                break

        # Check if the column was found
        if column_index is not None:
            print(f"The column index for '{target_date}' is: {column_index}")
        else:
            print(f"Column '{target_date}' not found in the sheet.")
    return column_indices


def fill_radio_voice_columns(region_wise_radio_voice_sheet, voice_regional_pivot_sheet, column_indices):
    #region-wise-radio-Voice
    for i in region_wise_radio_voice_sheet.iter_rows():  # iter_rows() --> 0 based index

        sales = i[1].value  # heading row = 1, Sales = 1
        row_number = i[1].row

        for j in voice_regional_pivot_sheet.iter_rows():  # min_row=3
            if sales is not None and j[0].value is not None and j[0].value.lower() == sales.lower():
                for k in range(len(column_indices)):
                    region_wise_radio_voice_sheet.cell(row=row_number, column=column_indices[k]).value = j[k+1].value


def fill_radio_voice_blank_with_na(region_wise_radio_voice_sheet, column_indices):
    # sei ghor gula #N/A or None segulare 0 diye fillup kortesi
    for row in region_wise_radio_voice_sheet.iter_rows():
        sales = row[1].value
        row_number = row[1].row

        if sales is None or sales == ' ' or sales == '':
            continue

        for k in range(len(column_indices)):  # 5, 6
            if region_wise_radio_voice_sheet.cell(row=row_number, column=column_indices[k]).value is None:
                region_wise_radio_voice_sheet.cell(row=row_number, column=column_indices[k]).value = 0


def fill_radio_data_columns(region_wise_radio_data_sheet, data_regional_pivot_sheet, column_indices):
    # region-wise-radio-Data
    for i in region_wise_radio_data_sheet.iter_rows():  # iter_rows() --> 0 based index

        sales = i[1].value  # heading row = 1, Sales = 1
        row_number = i[1].row

        for j in data_regional_pivot_sheet.iter_rows():
            if sales is not None and j[0].value is not None and j[0].value.lower() == sales.lower():
                for k in range(len(column_indices)):
                    region_wise_radio_data_sheet.cell(row=row_number, column=column_indices[k]).value = j[k+1].value


def fill_radio_data_blank_with_na(region_wise_radio_data_sheet, column_indices):
    # sei ghor gula #N/A or None segulare 0 diye fillup kortesi
    for row in region_wise_radio_data_sheet.iter_rows():
        sales = row[1].value
        row_number = row[1].row

        if sales is None or sales == ' ' or sales == '':
            continue

        for k in range(len(column_indices)):  # 5, 6
            if region_wise_radio_data_sheet.cell(row=row_number, column=column_indices[k]).value is None:
                region_wise_radio_data_sheet.cell(row=row_number, column=column_indices[k]).value = 0

    column_indices.clear()


def save_file(technology_region_wb):
    technology_region_wb.save(daily_technology_file_name)


def region_wise_radio_sheet_processing(assign_from_date, assign_to_date):
    technology_region_wb, regional_wb = load_workbook(daily_technology_file_name, regional_file_name)
    region_wise_radio_voice_sheet, region_wise_radio_data_sheet, voice_regional_pivot_sheet, data_regional_pivot_sheet = load_worksheet(technology_region_wb, regional_wb)
    current_year, formatted_date_list = insert_dates_to_list(assign_from_date, assign_to_date)
    formatted_dates = format_inserted_dates(current_year, formatted_date_list)
    column_indices = get_column_indices(region_wise_radio_voice_sheet, formatted_dates)
    fill_radio_voice_columns(region_wise_radio_voice_sheet, voice_regional_pivot_sheet, column_indices)
    fill_radio_voice_blank_with_na(region_wise_radio_voice_sheet, column_indices)
    fill_radio_data_columns(region_wise_radio_data_sheet, data_regional_pivot_sheet, column_indices)
    fill_radio_data_blank_with_na(region_wise_radio_data_sheet, column_indices)
    save_file(technology_region_wb)