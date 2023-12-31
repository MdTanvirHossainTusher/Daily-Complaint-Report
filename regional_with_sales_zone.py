import os
import openpyxl
import pandas as pd
from openpyxl.styles import Font
import tkinter as tk
from tkinter import filedialog
from tkcalendar import DateEntry
import tech_complaint_sheet as tech_complaint
import regional_file_pivot_table as voice_data_pivot
import region_wise_radio_sheet as region_wise


regional_file_name = r'regional.xlsx'
final_regional_file_name = r'regional_final.xlsx'
site_to_cc_file_name_csv = r'Site to CC.csv'
site_to_cc_file_name_excel = r'Site to CC.xlsx'
on_air_site_file_name = r'ON AIR SITES DETAILS.xlsx'
intermediate_regional_file_name = r'Intermediate regional file.xlsx'
on_air_site_sheet_name = 'Sites'

open_from_date = ""
open_to_date = ""
assign_from_date = ""
assign_to_date = ""

site_id_col_index = 19  # 1 based index
district_col_index = 20  # 1 based index
tf_col_index = 21  # 1 based index
sales_col_index = 22  # 1 based index

site_id_col_name = 'Site id'
district_col_name = 'District'
tf_col_name = 'T/F'
sales_col_name = 'Sales'

site_id_heading_cell = 'S1'
district_heading_cell = 'T1'
tf_heading_cell = 'U1'
sales_heading_cell = 'V1'


def load_workbook_sheet(regional_file_name):
    regional_wb = openpyxl.load_workbook(regional_file_name)
    regional_sheet = regional_wb.active
    return regional_wb, regional_sheet


def csv_to_excel():
    read_file = pd.read_csv(site_to_cc_file_name_csv)
    read_file.to_excel(site_to_cc_file_name_excel, index=None, header=True)
    return read_file


def load_site_to_cc_workbook_sheet():
    site_to_cc_wb = openpyxl.load_workbook(site_to_cc_file_name_excel, data_only=True)
    site_to_cc_sheet = site_to_cc_wb.active
    return site_to_cc_wb, site_to_cc_sheet


def load_on_air_site_workbook_sheet():
    # ON AIR SITE file contains some formula in it. That's why "data_only=True" attribute required
    on_air_site_wb = openpyxl.load_workbook(on_air_site_file_name, data_only=True)
    on_air_site_sheet = on_air_site_wb[on_air_site_sheet_name]
    return on_air_site_wb, on_air_site_sheet


def insert_columns(regional_sheet):
    regional_sheet.insert_cols(site_id_col_index)  # 1 based index
    regional_sheet[site_id_heading_cell] = site_id_col_name
    regional_sheet[site_id_heading_cell].font = Font(bold=True)

    regional_sheet.insert_cols(district_col_index)  # 1 based index
    regional_sheet[district_heading_cell] = district_col_name
    regional_sheet[district_heading_cell].font = Font(bold=True)

    regional_sheet.insert_cols(tf_col_index)  # 1 based index
    regional_sheet[tf_heading_cell] = tf_col_name
    regional_sheet[tf_heading_cell].font = Font(bold=True)

    regional_sheet.insert_cols(sales_col_index)  # 1 based index
    regional_sheet[sales_heading_cell] = sales_col_name
    regional_sheet[sales_heading_cell].font = Font(bold=True)

    print("columns inserted...\n")

    return regional_sheet


def fill_site_id_column(regional_sheet, site_to_cc_sheet):
    # site id col fill up
    heading_row = 1
    for i in regional_sheet.iter_rows():  # iter_rows() --> 0 based index
        if heading_row == 1:
            heading_row = 2
            continue

        id = i[0].value  # 7 = Team, 8 = sub category, heading row = 1
        row_number = i[0].row

        for j in site_to_cc_sheet.iter_rows():
            origin = j[0].value
            if id == origin:
                regional_sheet.cell(row=row_number, column=19).value = j[1].value  # column=8 = Team


def fill_district_sales_columns(regional_sheet, on_air_site_sheet):
    # district & sales column fill up
    heading_row = 1
    for i in regional_sheet.iter_rows():  # iter_rows() --> 0 based index
        if heading_row == 1:
            heading_row = 2
            continue

        district = i[18].value  # site id = 19
        row_number = i[18].row

        for j in on_air_site_sheet.iter_rows():
            generic_id = j[1].value

            if district == generic_id:
                regional_sheet.cell(row=row_number, column=20).value = j[31].value   # j[31] = 0 based, 31 e hbe cz manually col A bad diye dhortam.
                regional_sheet.cell(row=row_number, column=22).value = j[34].value   # j[34] = 0 based


def fill_empty_cell_with_na(regional_sheet):
    # empty cell fill up with #N/A value
    for rows in regional_sheet.iter_rows():
        if rows[19].value is None or rows[19].value == ' ' or rows[19].value == '':  # district = 31(here, 0 based)
            rows[19].value = "#N/A"
        if rows[21].value is None or rows[21].value == ' ' or rows[21].value == '':
            rows[21].value = "#N/A"


def fill_tf_column(regional_wb, regional_sheet):

    for rows in regional_sheet.iter_rows(min_row=1):
        if rows[13].value is None:
            continue
        elif rows[13].value.strip().lower() == '#N/A'.lower() or rows[19].value.strip().lower() == '#N/A'.lower():
            rows[20].value = '#N/A'
        elif rows[13].value.strip().lower() == rows[19].value.strip().lower():
            rows[20].value = 'TRUE'
        elif rows[13].value.strip().lower() != rows[19].value.strip().lower():
            rows[20].value = 'FALSE'
        elif rows[13].value.strip().lower() is None or rows[13].value.strip().lower() == ' ' or rows[13].value.strip().lower() == '':
            rows[20].value = 'TRUE'

    print("Newly added columns filled...\n")
    regional_wb.save(r'regional with filled columns.xlsx')


def are_strings_similar(str1, str2, max_allowed_diff=2):
    if len(str1) != len(str2):
        return False
    diff_count = sum(c1 != c2 for c1, c2 in zip(str1, str2))
    return diff_count <= max_allowed_diff


def clean_tf_column(regional_sheet):
    for rows in regional_sheet.iter_rows():
        district = rows[13].value
        calculated_district = rows[19].value

        if district == 'Chapai Nawabganj' and calculated_district == 'Nawabganj':
            rows[13].value = rows[19].value
            rows[20].value = 'TRUE'
        elif len(district) != len(calculated_district):
            rows[20].value = 'FALSE'
        elif district is not None and calculated_district is not None and are_strings_similar(district, calculated_district, max_allowed_diff=2):
            rows[13].value = rows[19].value
            rows[20].value = 'TRUE'


def remove_false_na_tf_values(regional_sheet, intermediate_regional_file_name):

    new_workbook = openpyxl.Workbook()
    new_workbook.save(intermediate_regional_file_name)

    regional_intermediate_wb = openpyxl.load_workbook(intermediate_regional_file_name, data_only=True)
    regional_intermediate_sheet = regional_intermediate_wb.active

    for row in regional_sheet.iter_rows():
        tf_col = row[20].value
        if tf_col is not None and tf_col.strip().lower() == 'true':
            regional_intermediate_sheet.append([cell.value for cell in row])

    return regional_intermediate_wb, regional_intermediate_sheet


def save_file_with_district_tf_columns(regional_intermediate_wb, intermediate_regional_file_name):
    regional_intermediate_wb.save(intermediate_regional_file_name)


def delete_calculated_district_tf_cols(regional_intermediate_sheet):
    columns_to_delete = [20, 21]  # district, T/F
    # Iterate over the columns to be deleted in reverse order
    for col in reversed(sorted(columns_to_delete)):
        regional_intermediate_sheet.delete_cols(col)


def save_final_regional_file(regional_intermediate_wb):
    regional_intermediate_wb.save(final_regional_file_name)


def browse_file():
    global regional_file_name
    regional_file_name = filedialog.askopenfilename(
        title="Select Regional File",
        filetypes=[("Excel Files", "*.xlsx;*.csv"), ("CSV files", "*.csv")]
    )
    print(f"Selected File: {regional_file_name}")


def browse_dates():
    global assign_from_date, assign_to_date, open_from_date, open_to_date

    open_from_date = entry_from_date.get_date().strftime("%Y-%m-%d")  # "2023-12-12" --> year-month-date
    open_to_date = entry_to_date.get_date().strftime("%Y-%m-%d")

    # Convert assign dates to the "12-DEC-23" format

    assign_from_date = entry_from_date.get_date().strftime("%d-%b-%y")  # "12-DEC-23"
    assign_to_date = entry_to_date.get_date().strftime("%d-%b-%y")

    print(f"Open From Date: {open_from_date}")
    print(f"Open To Date: {open_to_date}")
    print(f"Assign From Date: {assign_from_date}")
    print(f"Assign To Date: {assign_to_date}")


def get_path_directory():
    directory_name = os.path.dirname(regional_file_name)
    return directory_name.replace('/', '\\')


def regional_file_processing():
    regional_wb, regional_sheet = load_workbook_sheet(regional_file_name)
    read_file = csv_to_excel()
    site_to_cc_wb, site_to_cc_sheet = load_site_to_cc_workbook_sheet()
    on_air_site_wb, on_air_site_sheet = load_on_air_site_workbook_sheet()
    regional_sheet = insert_columns(regional_sheet)
    fill_site_id_column(regional_sheet, site_to_cc_sheet)
    fill_district_sales_columns(regional_sheet, on_air_site_sheet)
    fill_empty_cell_with_na(regional_sheet)
    fill_tf_column(regional_wb, regional_sheet)
    clean_tf_column(regional_sheet)
    regional_intermediate_wb, regional_intermediate_sheet = remove_false_na_tf_values(regional_sheet, intermediate_regional_file_name)
    save_file_with_district_tf_columns(regional_intermediate_wb, intermediate_regional_file_name)
    delete_calculated_district_tf_cols(regional_intermediate_sheet)
    save_final_regional_file(regional_intermediate_wb)


def run_process():

    path_directory = get_path_directory()
    regional_file_processing()
    print("saving8...")
    tech_complaint.daily_technology_tech_complaint_sheet_processing(assign_from_date, assign_to_date)
    print("saving9...")
    voice_data_pivot.regional_voice_data_pivot_table_create(path_directory)
    print("saving10...")
    region_wise.region_wise_radio_sheet_processing(assign_from_date, assign_to_date)
    print("saving11...")

    root.destroy()  # Close the GUI window


root = tk.Tk()
root.title("Regional file processing")

root.geometry("400x300")  # width and height

# File Selection
btn_browse_file = tk.Button(root, text="Browse Regional File", command=browse_file)
btn_browse_file.pack(pady=10)

# Date Selection using tkcalendar
entry_from_date = DateEntry(root, width=15, justify="center")
entry_to_date = DateEntry(root, width=15, justify="center")

# Set the date pattern directly on the DateEntry widget
entry_from_date.date_pattern = "%Y-%m-%d"  # "2023-12-12" --> year-month-date
entry_to_date.date_pattern = "%Y-%m-%d"

entry_from_date.pack(pady=5)
entry_to_date.pack(pady=5)

btn_browse_dates = tk.Button(root, text="Apply Dates", command=browse_dates)
btn_browse_dates.pack(pady=10)

# Run Process Button
btn_run_process = tk.Button(root, text="Run Process", command=run_process)
btn_run_process.pack(pady=20)

root.mainloop()
