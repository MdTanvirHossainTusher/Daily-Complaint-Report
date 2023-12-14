import openpyxl


daily_dump_file_name_raw_dump = r'Daily_Dump(Updated)-13 DEC-check.xlsx'
team_col_index = 8  # 1 based index
hash_na = '#N/A'

# ms excel/jupyter notebook
open_from_date = "2023-12-12"  # year-date-month
open_to_date = "2023-12-12"
assign_from_date = "12-DEC-23"
assign_to_date = "12-DEC-23"

# open_from_date = "12/11/2023"  # month/date/year
# open_to_date = "12/11/2023"
# assign_from_date = "11-Dec-23"
# assign_to_date = "11-Dec-23"
open_date_col_index = 20  # 0 based index
assign_date_col_index = 21  # 0 based index

open_date_filtered_rows = []
assign_date_filtered_rows = []

output_file = r'Raw Dump.xlsx'


def load_workbook():
    daily_dump_to_raw_dump_copy_data_workbook = openpyxl.load_workbook(daily_dump_file_name_raw_dump)
    return daily_dump_to_raw_dump_copy_data_workbook


def load_worksheet(daily_dump):
    daily_dump_sheet = daily_dump.active
    return daily_dump_sheet

############################### RAW & Assigned Sheet ###############################

## RAW Sheet
def create_raw_dump_file():
    raw_dump_workbook = openpyxl.Workbook()
    default_sheet = raw_dump_workbook['Sheet']
    raw_dump_workbook.remove(default_sheet)
    raw_sheet = raw_dump_workbook.create_sheet("RAW")
    assigned_sheet = raw_dump_workbook.create_sheet("Assigned")
    # output_file = 'Raw Dump.xlsx'
    raw_dump_workbook.save(output_file)
    return raw_dump_workbook, raw_sheet, assigned_sheet


def select_prv_days_from_open_date_col(daily_dump_sheet):
    heading_row = next(
        daily_dump_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    open_date_filtered_rows = [heading_row]

    for row in daily_dump_sheet.iter_rows(min_row=1, values_only=True):  # tuple = iter_rows() --> 0 based index

        team_value = row[team_col_index - 1]  # 7 = Team
        open_date_value = row[open_date_col_index]  # open date = 20

        if team_value != hash_na and open_from_date <= open_date_value <= open_from_date:
            open_date_filtered_rows.append(row)

    return open_date_filtered_rows


def paste_filtered_data_to_raw_sheet(raw_sheet_of_raw_dump, open_date_filtered_rows):
    for row in open_date_filtered_rows:
        raw_sheet_of_raw_dump.append(row)


def select_prv_assigned_date(daily_dump_sheet):
    heading_row = next(
        daily_dump_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assign_date_filtered_rows = [heading_row]

    for row in daily_dump_sheet.iter_rows(min_row=1, values_only=True):  # tuple = iter_rows() --> 0 based index

        team_value = row[team_col_index - 1]  # 7 = Team
        assign_date_value = row[assign_date_col_index]  # assign date = 21

        if team_value != hash_na and assign_date_value is not None and assign_from_date <= assign_date_value <= assign_to_date:
            assign_date_filtered_rows.append(row)

    return assign_date_filtered_rows


def paste_assigned_date_to_assigned_sheet(assigned_sheet_of_raw_dump, assign_date_filtered_rows):
    for row_data in assign_date_filtered_rows:
        assigned_sheet_of_raw_dump.append(row_data)


def save_daily_dump(raw_dump):
    raw_dump.save(output_file)


if __name__ == '__main__':
    daily_dump = load_workbook()
    daily_dump_sheet = load_worksheet(daily_dump)
    raw_dump_workbook, raw_sheet, assigned_sheet = create_raw_dump_file()

    open_date_filtered_rows = select_prv_days_from_open_date_col(daily_dump_sheet)
    paste_filtered_data_to_raw_sheet(raw_sheet, open_date_filtered_rows)

    assign_date_filtered_rows = select_prv_assigned_date(daily_dump_sheet)
    paste_assigned_date_to_assigned_sheet(assigned_sheet, assign_date_filtered_rows)

    save_daily_dump(raw_dump_workbook)

    open_date_filtered_rows.clear()
    assign_date_filtered_rows.clear()