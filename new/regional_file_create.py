import openpyxl


daily_dump_file_name_raw_dump = r'Daily_Dump(Updated)-13 DEC-check.xlsx'
team_col_index = 8  # 1 based index
hash_na = '#N/A'
output_file = r'13 Dec regional.xlsx'
assign_from_date = "12-Dec-23"
assign_to_date = "12-Dec-23"
open_date_col_index = 20  # 0 based index
assign_date_col_index = 21  # 0 based index

assign_date_filtered_rows = []


def load_workbook():
    daily_dump_to_raw_dump_copy_data_workbook = openpyxl.load_workbook(daily_dump_file_name_raw_dump)
    return daily_dump_to_raw_dump_copy_data_workbook


def load_worksheet(daily_dump):
    daily_dump_sheet = daily_dump.active
    return daily_dump_sheet


def create_regional_file():
    regional_workbook = openpyxl.Workbook()
    default_sheet = regional_workbook['Sheet']
    regional_workbook.remove(default_sheet)
    regional_sheet = regional_workbook.create_sheet("Sheet1")
    # output_file = '13 Dec regional.xlsx'
    regional_workbook.save(output_file)
    return regional_workbook, regional_sheet


def select_prv_assigned_date(daily_dump_sheet):
    heading_row = next(
        daily_dump_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assign_date_filtered_rows = [heading_row]

    for row in daily_dump_sheet.iter_rows(min_row=1, values_only=True):  # tuple = iter_rows() --> 0 based index

        team_value = row[team_col_index - 1]  # 7 = Team
        assign_date_value = row[assign_date_col_index]  # assign date = 21
        # print(assign_date_value, end=' ----\n')
        if team_value != hash_na and assign_date_value is not None and str(assign_from_date).lower() <= str(assign_date_value).lower() <= str(assign_to_date).lower():
            assign_date_filtered_rows.append(row)

    return assign_date_filtered_rows


def paste_assigned_date_to_assigned_sheet(assigned_sheet_of_raw_dump, assign_date_filtered_rows):
    for row_data in assign_date_filtered_rows:
        assigned_sheet_of_raw_dump.append(row_data)


def save_regional_file(regional):
    regional.save(output_file)


if __name__ == '__main__':
    daily_dump = load_workbook()
    daily_dump_sheet = load_worksheet(daily_dump)
    regional_workbook, regional_sheet = create_regional_file()

    assign_date_filtered_rows = select_prv_assigned_date(daily_dump_sheet)
    paste_assigned_date_to_assigned_sheet(regional_sheet, assign_date_filtered_rows)

    save_regional_file(regional_workbook)

    assign_date_filtered_rows.clear()