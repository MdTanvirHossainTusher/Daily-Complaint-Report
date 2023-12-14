import win32com.client
import openpyxl


# regional_file_name = r'13 Dec regional final.xlsx'
# regional_file_name_with_pivot_table_sheet = r'13 Dec regional pivot.xlsx'
# regional_file_with_pivot_table = r'I:\Openpyxl_tutorial\test\13 Dec regional final with pivot table.xlsx'


def insert_pt_field_set1(pt):

    field_filters = {}
    field_filters['sub_category'] = pt.PivotFields("SUB_CATEGORY")

    field_columns = {}
    field_columns['assign_date'] = pt.PivotFields("ASSIGNED_DATE")

    field_rows = {}
    field_rows['sales'] = pt.PivotFields("sales")

    field_values = {}
    field_values['sales_count'] = pt.PivotFields("sales")

    field_filters['sub_category'].Orientation = 3  # hidden = 0, row = 1, column = 2, page = 3, data = 4,

    field_rows['sales'].Orientation = 1  # hidden = 0, row = 1, column = 2, page = 3, data = 4,

    field_columns['assign_date'].Orientation = 2

    field_values['sales_count'].Orientation = 4
    field_values['sales_count'].Function = -4112  # count = -4112
    field_values['sales_count'].NumberFormat = "#,##0"


def pivot_table_creation_all(workbook, ws_data, ws_report, output_starting_cell, pivot_table_name):

    pt_cache = workbook.PivotCaches().Create(1, ws_data.Range("A1").CurrentRegion)
    pt = pt_cache.CreatePivotTable(ws_report.Range(output_starting_cell), pivot_table_name)
    pt.ColumnGrand = True
    pt.RowGrand = False
    pt.TableStyle2 = "PivotStyleMedium9"
    insert_pt_field_set1(pt)

    pivot_table = ws_report.PivotTables(pivot_table_name)
    pivot_field_product = pivot_table.PivotFields("SUB_CATEGORY")
    return pivot_field_product


def filter_multiple_items(workbook, ws_data, ws_report, output_starting_cell, pivot_table_name, items_to_exclude):

    pivot_field_product = pivot_table_creation_all(workbook, ws_data, ws_report, output_starting_cell, pivot_table_name)
    pivot_field_product.ClearAllFilters()
    pivot_field_product.EnableMultiplePageItems = True

    for item_name in items_to_exclude:
        try:
            pivot_field_product.PivotItems(item_name).Visible = False
        except Exception as e:
            print(f"Item '{item_name}' not found in the pivot table. Skipping...")


def filter_item():

    open_workbook = openpyxl.load_workbook(r'13 Dec regional final.xlsx')

    voice_sheet_name = "voice_pivot"
    open_workbook.create_sheet(title=voice_sheet_name)
    data_sheet_name = "data_pivot"
    open_workbook.create_sheet(title=data_sheet_name)
    open_workbook.save(r'13 Dec regional pivot.xlsx')

    # pywin
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True  # Optional: Set to True if you want to see Excel while the code is running

    workbook = excel.Workbooks.Open(r'I:\Openpyxl_tutorial\test\13 Dec regional pivot.xlsx')
    ws_data = workbook.Worksheets("Sheet1")

    ws_voice_pivot = workbook.Sheets(voice_sheet_name)
    ws_data_pivot = workbook.Sheets(data_sheet_name)

    # Voice
    output_starting_cell = "A3"
    pivot_table_name = "PivotTable1"
    items_to_exclude = ["Can't browse internet", "Data Speed Complaint"]
    filter_multiple_items(workbook, ws_data, ws_voice_pivot, output_starting_cell, pivot_table_name, items_to_exclude)

    # Data
    output_starting_cell = "A3"
    pivot_table_name = "PivotTable1"
    items_to_exclude = ["BAD VOICE QUALITY", "Call Drop", "Coverage Complaint", "MULTIPLE RETRIES"]
    filter_multiple_items(workbook, ws_data, ws_data_pivot, output_starting_cell, pivot_table_name, items_to_exclude)

    workbook.SaveAs(r'I:\Openpyxl_tutorial\test\13 Dec regional final with pivot table.xlsx')
    workbook.Close()

    excel.Quit()


if __name__ == "__main__":
    filter_item()
