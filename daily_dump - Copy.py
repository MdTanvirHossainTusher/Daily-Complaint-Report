import openpyxl
import pandas as pd
from openpyxl.styles import Font

# daily_dump_file_name_csv = r'Daily_Dump(Updated).csv'
daily_dump_file_name_csv = r'Daily_Dump(Updated)-12 DEC.csv'
daily_dump_file_name_excel = r'Daily_Dump(Updated)-12 DEC.xlsx'
category_team_file_name = r'Category team.xlsx'
team_col_index = 8  # 1 based index
team_col_name = 'Team'
team_col_heading = 'H1'
sub_category_col_num = 8  # 0 based index
hash_na = '#N/A'
daily_dump_id_col_index = 0  # 0 based index
escalation_virtual_team_index = 2  # 0 based index
closing_parenthesis_sign = ')'
vas_team_name = 'VAS'


def keep_ascii_printable(text):
    if pd.isna(text):  # Check if the cell is NaN (blank)
        return text
    return ''.join(char for char in str(text) if 32 <= ord(char) <= 126)


def csv_to_excel_conversion():
    read_file = pd.read_csv(daily_dump_file_name_csv, dtype=str, low_memory=False)
    read_file = read_file.apply(lambda x: x.map(keep_ascii_printable))
    read_file.to_excel(daily_dump_file_name_excel, index=None, header=True)


def load_workbook():
    daily_dump = openpyxl.load_workbook(daily_dump_file_name_excel)
    category_team = openpyxl.load_workbook(category_team_file_name)
    return daily_dump, category_team

# raw_dump = openpyxl.load_workbook(r'RAW Dump for daily trending - Copy.xlsx')
# print("done!")


def load_worksheet(daily_dump, category_team):
    daily_dump_sheet = daily_dump.active
    category_team_sheet = category_team.active
    return daily_dump_sheet, category_team_sheet

# raw_sheet_of_raw_dump = raw_dump['RAW']
# assigned_sheet_of_raw_dump = raw_dump['Assigned']
# raw_pivot_sheet_of_raw_dump = raw_dump['RAW Pivot']
# assigned_pivot_sheet_of_raw_dump = raw_dump['Assigned Pivot']

# print("done!!!")


def insert_team_col(daily_dump_sheet):
    daily_dump_sheet.insert_cols(team_col_index)  # 1 based index
    daily_dump_sheet[team_col_heading] = team_col_name
    daily_dump_sheet[team_col_heading].font = Font(bold=True)

# heading_row = 1
# cnt = 1


def fill_team_col_from_team_category_file(daily_dump_sheet, category_team_sheet):
    # fill up Team column using vlookup between daily_dump and category_team file
    heading_row = 1
    for i in daily_dump_sheet.iter_rows():  # iter_rows() --> 0 based index
        if heading_row == 1:
            heading_row = 2
            continue

        sub_category = i[sub_category_col_num].value  # 8 = sub category (0-based)
        row_number = i[sub_category_col_num].row

        for j in category_team_sheet.iter_rows():
            if sub_category is not None and j[daily_dump_id_col_index].value is not None \
                    and j[daily_dump_id_col_index].value.strip().lower() == sub_category.strip().lower():
                daily_dump_sheet.cell(row=row_number, column=team_col_index).value = j[escalation_virtual_team_index].value  # column = 8 = Team


def fill_blank_team_cell_with_na(daily_dump_sheet):
    # fill up blank cell using '#N/A' in Team column
    for row in daily_dump_sheet.iter_rows():
        if row[team_col_index - 1].value is None:  # 7 = Team, team_col_index = 8
            row[team_col_index - 1].value = hash_na


def team_with_short_coded_subcategory(daily_dump_sheet):
    # Select those `#N/A` from `Team` col where sub_category has `short code` in it.
    for i in daily_dump_sheet.iter_rows():  # tuple = iter_rows() --> 0 based index

        team_value = i[team_col_index - 1].value  # 7 = Team
        sub_category_value = i[sub_category_col_num].value  # 8 = sub-category
        row_number = i[team_col_index - 1].row

        if team_value == hash_na and sub_category_value[-1] == closing_parenthesis_sign and sub_category_value[-2].isnumeric():
            daily_dump_sheet.cell(row=row_number, column=team_col_index).value = vas_team_name  # column = 8 = Team


def save_daily_dump(daily_dump):
    # daily_dump.save(daily_dump_file_name_excel)
    daily_dump.save(r'Daily_Dump(Updated)-12 DEC-check.xlsx')


if __name__ == '__main__':
    csv_to_excel_conversion()
    daily_dump, category_team = load_workbook()
    daily_dump_sheet, category_team_sheet = load_worksheet(daily_dump, category_team)
    insert_team_col(daily_dump_sheet)
    fill_team_col_from_team_category_file(daily_dump_sheet, category_team_sheet)
    fill_blank_team_cell_with_na(daily_dump_sheet)
    team_with_short_coded_subcategory(daily_dump_sheet)
    save_daily_dump(daily_dump)


'''

############################### RAW & Assigned Sheet ###############################

## RAW Sheet

from_date = "10/7/2023"
to_date = "10/8/2023"
cnt = 1

open_date_filtered_rows = []

def select_prv_days_from_open_date_col(cnt):

    for row in daily_dump_sheet.iter_rows(min_row=2, values_only=True):  # tuple = iter_rows() --> 0 based index

        team_value = row[7]  # 7 = Team, 8 = sub category, heading row = 1
        open_date_value = row[20]  # open date = 20

        if team_value != hash_na and from_date <= open_date_value <= to_date:
            open_date_filtered_rows.append(row)
    #         cnt += 1
    #         print(team_value, open_date_value)
    # print(cnt)


select_prv_days_from_open_date_col(1)


def paste_filtered_data_to_raw_sheet(count, c):

    for row in raw_sheet_of_raw_dump.iter_rows():
        # count += 1
        if row[0].value is None:
            for row_data in open_date_filtered_rows:
                # c += 1
                raw_sheet_of_raw_dump.append(row_data)

            break
    # print(count, c)
    # raw_dump.save(r'RAW Dump for daily trending - Copy2.xlsx')


paste_filtered_data_to_raw_sheet(0, 0)



## Assigned Sheet

assign_from_date = "7-Oct-23"
assign_to_date = "8-Oct-23"
assign_date_filtered_rows = []


def select_prv_assigned_date(cnt):
    for row in daily_dump_sheet.iter_rows(min_row=2, values_only=True):  # tuple = iter_rows() --> 0 based index

        team_value = row[7]  # 7 = Team, 8 = sub category, heading row = 1
        assign_date_value = row[21]  # assign date = 21

        if team_value != '#N/A' and assign_date_value is not None and assign_from_date <= assign_date_value <= assign_to_date:
            assign_date_filtered_rows.append(row)
            # cnt += 1
    #         print(team_value, assign_date_value)
    # print(cnt, len(assign_date_filtered_rows), end=' +++\n ')


select_prv_assigned_date(0)



def paste_assigned_date_to_assigned_sheet(count, c):

    for row_data in assign_date_filtered_rows:
        c += 1
        assigned_sheet_of_raw_dump.append(row_data)
    print(count, c)
    raw_dump.save(r'RAW Dump for daily trending - Copy2.xlsx')

paste_assigned_date_to_assigned_sheet(0, 0)
'''

'''

############# radio dump #######################

def select_prv_assigned_date(cnt):
    heading = True
    for row in daily_dump_sheet.iter_rows(min_row=1, values_only=True):  # tuple = iter_rows() --> 0 based index

        team_value = row[7]  # 7 = Team, 8 = sub category, heading row = 1
        assign_date_value = row[21]  # assign date = 21

        if heading:
            assign_date_filtered_rows.append(row)
            heading = False
            # print(team_value, assign_date_value)

        if team_value != '#N/A' and assign_date_value is not None and assign_from_date <= assign_date_value <= assign_to_date:
            assign_date_filtered_rows.append(row)
            # cnt += 1
            # print(team_value, assign_date_value)
    # print(cnt, len(assign_date_filtered_rows), end=' +++\n ')


select_prv_assigned_date(0)

def paste_to_radio_dump(c):

    radio_workbook = openpyxl.Workbook()
    radio_worksheet = radio_workbook.active
    # radio_worksheet.append(daily_dump_sheet[1])

    for row_data in assign_date_filtered_rows:
        # c += 1
        radio_worksheet.append(row_data)
    # print(c)

    radio_worksheet.insert_cols(19)  # 1 based index
    radio_worksheet.insert_cols(20)  # 1 based index
    radio_worksheet['S1'] = 'site id'
    radio_worksheet['T1'] = 'sales'

    radio_dump_file = "Radio Dump.xlsx"
    radio_workbook.save(radio_dump_file)

paste_to_radio_dump(0)

'''
