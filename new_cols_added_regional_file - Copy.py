import openpyxl
import pandas as pd
from openpyxl.styles import Font


# Load 'regional.xlsx'
regional_wb = openpyxl.load_workbook(r'Dec 12 regional 2.xlsx')
regional_sheet = regional_wb.active


read_file = pd.read_csv(r'site to cc Dec 12.csv')
read_file.to_excel(r'site to cc Dec 12.xlsx', index=None, header=True)

# Load 'Site to CC.xlsx'
site_to_cc_wb = openpyxl.load_workbook(r'site to cc Dec 12.xlsx')
site_to_cc_sheet = site_to_cc_wb.active

print('checking..')
on_air_site = openpyxl.load_workbook(r'ON AIR SITES DETAILS _230930.xlsx')
print('checking............')
on_air_site_sheet = on_air_site['Sites']

regional_sheet.insert_cols(19)  # 1 based index
regional_sheet['S1'] = 'site id'
regional_sheet['S1'].font = Font(bold=True)

regional_sheet.insert_cols(20)  # 1 based index
regional_sheet['T1'] = 'district'
regional_sheet['T1'].font = Font(bold=True)

regional_sheet.insert_cols(21)  # 1 based index
regional_sheet['U1'] = 'T/F'
regional_sheet['U1'].font = Font(bold=True)

regional_sheet.insert_cols(22)  # 1 based index
regional_sheet['V1'] = 'sales'
regional_sheet['V1'].font = Font(bold=True)


heading_row = 1

# site id col fill up
for i in regional_sheet.iter_rows():  # iter_rows() --> 0 based index
    if heading_row == 1:
        heading_row = 2
        continue

    id = i[0].value  # 7 = Team, 8 = sub category, heading row = 1
    row_number = i[0].row

    for j in site_to_cc_sheet.iter_rows():
        origin = j[0].value
        if id == origin:
            regional_sheet.cell(row=row_number, column=19).value = j[1].value  # column=8 = Team


# district & sales column fill up
for i in regional_sheet.iter_rows():  # iter_rows() --> 0 based index
    if heading_row == 1:
        heading_row = 2
        continue

    district = i[18].value  # site id = 19
    row_number = i[18].row

    for j in on_air_site_sheet.iter_rows():
        generic_id = j[1].value
        if district == generic_id:
            regional_sheet.cell(row=row_number, column=20).value = j[31].value   # j[31] = 0 based, 31 e hbe cz manually col A bad diye dhortam.
            regional_sheet.cell(row=row_number, column=22).value = j[34].value   # j[34] = 0 based

# empty cell fill up with #N/A value
for rows in regional_sheet.iter_rows():
    if rows[19].value is None or rows[19].value == ' ' or rows[19].value == '':  # district = 31(here, 0 based)
        rows[19].value = "#N/A"
    if rows[21].value is None or rows[21].value == ' ' or rows[21].value == '':
        rows[21].value = "#N/A"


# regional_wb = openpyxl.load_workbook('regional_updated4.xlsx')
# regional_sheet = regional_wb.active

for rows in regional_sheet.iter_rows():
    if rows[13].value is None:
        continue
    elif rows[13].value.strip().lower() == '#N/A'.lower() or rows[19].value.strip().lower() == '#N/A'.lower():
        rows[20].value = '#N/A'
    elif rows[13].value.strip().lower() == rows[19].value.strip().lower():
        rows[20].value = 'TRUE'
    elif rows[13].value.strip().lower() != rows[19].value.strip().lower():
        rows[20].value = 'FALSE'
    elif rows[13].value.strip().lower() is None or rows[13].value.strip().lower() == ' ' or rows[13].value.strip().lower() == '':
        rows[20].value = 'TRUE'


def are_strings_similar(str1, str2, max_allowed_diff=2):

    if len(str1) != len(str2):
        return False
    diff_count = sum(c1 != c2 for c1, c2 in zip(str1, str2))
    return diff_count <= max_allowed_diff


for rows in regional_sheet.iter_rows():
    district = rows[13].value
    calculated_district = rows[19].value

    if district == 'Chapai Nawabganj' and calculated_district == 'Nawabganj':
        rows[13].value = rows[19].value
        rows[20].value = 'TRUE'

    elif district is not None and calculated_district is not None and are_strings_similar(district, calculated_district, max_allowed_diff=2):
        rows[13].value = rows[19].value
        rows[20].value = 'TRUE'


for rows in regional_sheet.iter_rows():
    tf_col = rows[20].value
    if tf_col is not None and (tf_col.lower() == 'FALSE'.lower() or tf_col.lower() == '#N/A'.lower()):
        regional_sheet.delete_rows(rows[0].row)

regional_wb.save('regional_file_with_calculated_district_tf_cols_Dec_nn.xlsx')


def delete_calculated_district_tf_cols():
    columns_to_delete = [20, 21]

    # Iterate over the columns to be deleted in reverse order
    for col in reversed(sorted(columns_to_delete)):
        regional_sheet.delete_cols(col)


delete_calculated_district_tf_cols()

# Save the changes to 'regional.xlsx'
regional_wb.save('Dec 12 regional_nn.xlsx')

# Close the workbooks
# regional_wb.close()
# site_to_cc_wb.close()
