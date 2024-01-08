import openpyxl

regional_file_name = r'regional.xlsx'
output_file_name = r'Lat Long File.xls'


def load_regional_file(regional_file_path):
    return openpyxl.load_workbook(regional_file_path)


def load_regional_sheet(regional_workbook):
    return regional_workbook.active


def create_lat_long_file():
    lat_long_workbook = openpyxl.Workbook()
    lat_long_sheet = lat_long_workbook.active
    return lat_long_workbook, lat_long_sheet


def copy_data_to_lat_long_file(regional_sheet, lat_long_sheet):

    columns_to_copy = ['ID', 'GOOGLE_EARTH_LAT', 'GOOGLE_EARTH_LAT']
    new_columns = {'ID': 1, 'GOOGLE_EARTH_LAT': 31, 'GOOGLE_EARTH_LONG': 32}

    # Copy headers to the new file
    for col_num, col_name in enumerate(columns_to_copy, start=1):
        lat_long_sheet.cell(row=1, column=col_num, value=col_name)

    # Copy data to the new file
    for row_num in range(2, regional_sheet.max_row + 1):
        for col_num, col_name in enumerate(columns_to_copy, start=1):
            original_col_num = new_columns[col_name]
            value = regional_sheet.cell(row=row_num, column=original_col_num).value
            lat_long_sheet.cell(row=row_num, column=col_num, value=value)


def save_lat_long_file(lat_long_workbook):
    lat_long_file_path = output_file_name
    lat_long_workbook.save(lat_long_file_path)


def lat_long_file_creation_process():
    regional_workbook = load_regional_file(regional_file_name)
    regional_sheet = load_regional_sheet(regional_workbook)
    lat_long_workbook, lat_long_sheet = create_lat_long_file()
    copy_data_to_lat_long_file(regional_sheet, lat_long_sheet)
    save_lat_long_file(lat_long_workbook)