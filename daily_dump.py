import os
import openpyxl
import pandas as pd
from openpyxl.styles import Font
import tkinter as tk
from tkinter import filedialog
from tkcalendar import DateEntry
import create_raw as crd
import raw_pivot_table_filtering as raw_pivot
import assign_pivot_table_filtering as assign_pivot
import regional_file_create as create_regional
import lat_long_cleaning_process as clean_lat_long
import lat_long_file_code as lat_long


daily_dump_file_name_csv = ""
daily_dump_file_name_excel = r'Daily_Dump(Updated).xlsx'
category_team_file_name = r'Category team.xlsx'
team_col_index = 8  # 1 based index
team_col_name = 'Team'
team_col_heading = 'H1'
sub_category_col_num = 8  # 0 based index
hash_na = '#N/A'
daily_dump_id_col_index = 0  # 0 based index
escalation_virtual_team_index = 2  # 0 based index
closing_parenthesis_sign = ')'
vas_team_name = 'VAS'

open_from_date = ""
open_to_date = ""
assign_from_date = ""
assign_to_date = ""


def keep_ascii_printable(text):
    if pd.isna(text):  # Check if the cell is NaN (blank)
        return text
    return ''.join(char for char in str(text) if 32 <= ord(char) <= 126)


def csv_to_excel_conversion():
    read_file = pd.read_csv(daily_dump_file_name_csv, dtype=str, low_memory=False)
    read_file = read_file.apply(lambda x: x.map(keep_ascii_printable))
    read_file.to_excel(daily_dump_file_name_excel, index=None, header=True)
    print("converted...")


def load_workbook(daily_dump_file_name_excel):
    daily_dump = openpyxl.load_workbook(daily_dump_file_name_excel)
    category_team = openpyxl.load_workbook(category_team_file_name)

    print("both file loaded.\n")
    return daily_dump, category_team


def load_worksheet(daily_dump, category_team):
    daily_dump_sheet = daily_dump.active
    category_team_sheet = category_team.active
    return daily_dump_sheet, category_team_sheet


def insert_team_col(daily_dump_sheet):
    daily_dump_sheet.insert_cols(team_col_index)  # 1 based index
    daily_dump_sheet[team_col_heading] = team_col_name
    daily_dump_sheet[team_col_heading].font = Font(bold=True)


def fill_team_col_from_team_category_file(daily_dump_sheet, category_team_sheet):
    # fill up Team column using vlookup between daily_dump and category_team file
    heading_row = 1
    for i in daily_dump_sheet.iter_rows():  # iter_rows() --> 0 based index
        if heading_row == 1:
            heading_row = 2
            continue

        sub_category = i[sub_category_col_num].value  # 8 = sub category (0-based)
        row_number = i[sub_category_col_num].row

        for j in category_team_sheet.iter_rows():
            if sub_category is not None and j[daily_dump_id_col_index].value is not None \
                    and j[daily_dump_id_col_index].value.strip().lower() == sub_category.strip().lower():
                daily_dump_sheet.cell(row=row_number, column=team_col_index).value = j[escalation_virtual_team_index].value  # column = 8 = Team


def fill_blank_team_cell_with_na(daily_dump_sheet):
    # fill up blank cell using '#N/A' in Team column
    for row in daily_dump_sheet.iter_rows():
        if row[team_col_index - 1].value is None:  # 7 = Team, team_col_index = 8
            row[team_col_index - 1].value = hash_na


def team_with_short_coded_subcategory(daily_dump_sheet):
    # Select those `#N/A` from `Team` col where sub_category has `short code` in it.
    for i in daily_dump_sheet.iter_rows():  # tuple = iter_rows() --> 0 based index

        team_value = i[team_col_index - 1].value  # 7 = Team
        sub_category_value = i[sub_category_col_num].value  # 8 = sub-category
        row_number = i[team_col_index - 1].row

        if team_value == hash_na and sub_category_value[-1] == closing_parenthesis_sign and sub_category_value[-2].isnumeric():
            daily_dump_sheet.cell(row=row_number, column=team_col_index).value = vas_team_name  # column = 8 = Team


def save_daily_dump(daily_dump):
    daily_dump.save(daily_dump_file_name_excel)


def browse_file():
    global daily_dump_file_name_csv
    daily_dump_file_name_csv = filedialog.askopenfilename(
        title="Select Daily Dump File",
        filetypes=[("Excel Files", "*.xlsx;*.csv"), ("CSV files", "*.csv")]
    )
    print(f"Selected File: {daily_dump_file_name_csv}")


def browse_dates():
    global assign_from_date, assign_to_date, open_from_date, open_to_date

    open_from_date = entry_from_date.get_date().strftime("%Y-%m-%d")  # "2023-12-12" --> year-month-date
    open_to_date = entry_to_date.get_date().strftime("%Y-%m-%d")

    # Convert assign dates to the "12-DEC-23" format
    assign_from_date = entry_from_date.get_date().strftime("%d-%b-%y")  # "12-DEC-23"
    assign_to_date = entry_to_date.get_date().strftime("%d-%b-%y")

    print(f"Open From Date: {open_from_date}")
    print(f"Open To Date: {open_to_date}")
    print(f"Assign From Date: {assign_from_date}")
    print(f"Assign To Date: {assign_to_date}")


def get_path_directory():
    directory_name = os.path.dirname(daily_dump_file_name_csv)
    return directory_name.replace('/', '\\')


def daily_dump_processes():
    csv_to_excel_conversion()
    daily_dump, category_team = load_workbook(daily_dump_file_name_excel)
    daily_dump_sheet, category_team_sheet = load_worksheet(daily_dump, category_team)
    insert_team_col(daily_dump_sheet)
    fill_team_col_from_team_category_file(daily_dump_sheet, category_team_sheet)
    fill_blank_team_cell_with_na(daily_dump_sheet)
    team_with_short_coded_subcategory(daily_dump_sheet)
    save_daily_dump(daily_dump)


def run_process():
    daily_dump_processes()
    print("saving1...")
    crd.raw_dump_processes(open_from_date, open_to_date, assign_from_date, assign_to_date)
    print("saving2...")
    path_directory = get_path_directory()
    raw_pivot.filter_item(path_directory)
    print("saving3...")
    assign_pivot.filter_item(path_directory)
    print("saving4...")
    create_regional.regional_file_creation_process(assign_from_date, assign_to_date)
    print("saving5...")
    clean_lat_long.lat_long_cleaner()
    print("saving6...")
    lat_long.lat_long_file_creation_process()
    print("saving7...")
    #
    root.destroy()  # Close the GUI window


root = tk.Tk()
root.title("Daily Dump file processing")

root.geometry("400x300")  # width and height

# File Selection
btn_browse_file = tk.Button(root, text="Browse Daily Dump File", command=browse_file)
btn_browse_file.pack(pady=10)

# Date Selection using tkcalendar
entry_from_date = DateEntry(root, width=15, justify="center")
entry_to_date = DateEntry(root, width=15, justify="center")

# Set the date pattern directly on the DateEntry widget
entry_from_date.date_pattern = "%Y-%m-%d"  # "2023-12-12" --> year-month-date
entry_to_date.date_pattern = "%Y-%m-%d"

entry_from_date.pack(pady=5)
entry_to_date.pack(pady=5)

btn_browse_dates = tk.Button(root, text="Apply Dates", command=browse_dates)
btn_browse_dates.pack(pady=10)

# Run Process Button
btn_run_process = tk.Button(root, text="Run Process", command=run_process)
btn_run_process.pack(pady=20)

root.mainloop()


