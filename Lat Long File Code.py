import openpyxl

# Load the original Excel file
regional_file_path = 'I:\Openpyxl_tutorial\Parts\Dec_3_regional(NEW MINE).xlsx'
regional_workbook = openpyxl.load_workbook(regional_file_path)
regional_sheet = regional_workbook.active

# Create a new workbook for the 'Lat Long File.xls'
lat_long_workbook = openpyxl.Workbook()
lat_long_sheet = lat_long_workbook.active

# Define the columns to copy
columns_to_copy = ['ID', 'GOOGLE_EARTH_LAT', 'GOOGLE_EARTH_LONG']

# Map the new column positions in the new file
new_columns = {'ID': 1, 'GOOGLE_EARTH_LAT': 31, 'GOOGLE_EARTH_LONG': 32}

# Copy headers to the new file
for col_num, col_name in enumerate(columns_to_copy, start=1):
    lat_long_sheet.cell(row=1, column=col_num, value=col_name)

# Copy data to the new file
for row_num in range(2, regional_sheet.max_row + 1):
    for col_num, col_name in enumerate(columns_to_copy, start=1):
        original_col_num = new_columns[col_name]
        value = regional_sheet.cell(row=row_num, column=original_col_num).value
        lat_long_sheet.cell(row=row_num, column=col_num, value=value)

# Save the new workbook
lat_long_file_path = 'I:\Openpyxl_tutorial\Parts\Dec 3 Lat Long File (created)3.xls'
lat_long_workbook.save(lat_long_file_path)

# Close the workbooks
# original_workbook.close()
# new_workbook.close()
